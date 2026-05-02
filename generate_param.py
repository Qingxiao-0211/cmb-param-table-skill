#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
招商银行公募基金参数表自动填充脚本
功能：根据基金公司提供的标准参数表，自动填充招行参数表导入模板
"""

import openpyxl
import re
import os
import sys
from copy import copy
from datetime import datetime


def parse_standard_param_table(filepath):
    """解析标准参数表，提取所有字段"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    data = {}
    current_group = None

    for row in range(1, ws.max_row + 1):
        cell_1 = ws.cell(row=row, column=1).value
        cell_2 = ws.cell(row=row, column=2).value
        cell_3 = ws.cell(row=row, column=3).value
        cell_4 = ws.cell(row=row, column=4).value
        cell_5 = ws.cell(row=row, column=5).value

        if cell_1 is None:
            continue

        field_name = str(cell_1).strip()

        # 跳过标题行
        if '参数表' in field_name and len(field_name) > 10:
            continue

        # 处理有分组的情况（如基金认购费、申购资金清算账户等）
        if cell_3 is not None and cell_4 is not None and cell_5 is not None:
            group_name = str(cell_3).strip()
            sub_field = str(cell_4).strip()
            sub_value = cell_5

            if group_name not in data:
                data[group_name] = {}
            data[group_name][sub_field] = sub_value

        # 存储主字段
        if cell_2 is not None:
            data[field_name] = cell_2

    return data


def extract_holding_period(fund_short_name):
    """从基金简称中提取封闭期/持有期信息"""
    if not fund_short_name:
        return 0, 'M'

    fund_short_name = str(fund_short_name)
    # 匹配模式：数字+单位（月/季/年/日）
    patterns = [
        (r'(\d+)(个月|月)', 'M', '个月'),
        (r'(\d+)(日|天)', 'D', '日'),
        (r'(\d+)(年)', 'Y', '年'),
        (r'(\d+)(季)', 'S', '季'),
    ]

    for pattern, unit_code, unit_name in patterns:
        match = re.search(pattern, fund_short_name)
        if match:
            num = int(match.group(1))
            return num, unit_code

    return 0, 'M'


def extract_tn_value(tn_string):
    """从T+n或R+n字符串中提取数字"""
    if not tn_string:
        return None
    match = re.search(r'[T|R]\+(\d+)', str(tn_string))
    if match:
        return int(match.group(1))
    return None


def extract_date_range(date_range_str):
    """从日期范围字符串中提取开始和结束日期"""
    if not date_range_str:
        return None, None

    date_range_str = str(date_range_str)

    # 匹配模式：YYYY-MM-DD至YYYY-MM-DD
    pattern = r'(\d{4}-\d{2}-\d{2})[至\-～~](\d{4}-\d{2}-\d{2})'
    match = re.search(pattern, date_range_str)
    if match:
        return match.group(1), match.group(2)

    # 也可能只有一个日期
    pattern_single = r'(\d{4}-\d{2}-\d{2})'
    match = re.search(pattern_single, date_range_str)
    if match:
        return match.group(1), match.group(1)

    return None, None


def get_fund_type_mapping(fund_type_str, fund_name=''):
    """基金类型映射：将标准参数表的基金类型映射为招行参数表代码"""
    fund_type_str = str(fund_type_str).upper() if fund_type_str else ''
    fund_name = str(fund_name) if fund_name else ''

    # 根据基金名称关键词判断（更准确）
    if '货币' in fund_name or '现金' in fund_name:
        return 'M:货币基金'
    elif '债券' in fund_name:
        return 'B:债券基金'
    elif '股票' in fund_name and '指数' not in fund_name:
        return 'S:股票基金'

    # 根据标准参数表的"基金类型"字段判断
    if 'FOF' in fund_type_str or '基金中基金' in fund_type_str:
        return '1:混合基金'
    elif 'QDII' in fund_type_str:
        return 'W:全球'  # QDII默认全球，后续投资地区会细化
    elif '债券' in fund_type_str:
        return 'B:债券基金'
    elif '货币' in fund_type_str or '现金' in fund_type_str:
        return 'M:货币基金'
    elif '股票' in fund_type_str or '指数' in fund_type_str:
        return 'S:股票基金'
    elif '混合' in fund_type_str:
        return '1:混合基金'
    else:
        return '1:混合基金'  # 默认混合基金


def parse_fee_value(fee_str):
    """解析费率值，去掉百分号"""
    if not fee_str:
        return 0
    fee_str = str(fee_str).strip().replace('%', '')
    try:
        return float(fee_str)
    except:
        return 0


def get_max_fee(fee_dict):
    """从费率字典中获取最高档费率，排除固定金额门槛"""
    if not fee_dict or not isinstance(fee_dict, dict):
        return 0

    max_fee = 0
    for key, value in fee_dict.items():
        if not key or not value:
            continue
        val_str = str(value).strip()

        # 如果值包含百分号，是费率
        if '%' in val_str:
            fee_val = parse_fee_value(val_str)
            if fee_val > max_fee:
                max_fee = fee_val
        # 如果值是纯数字且小于10，认为是费率值（如0.5, 1.2等）
        elif val_str.replace('.', '').replace('-', '').isdigit():
            try:
                num_val = float(val_str)
                # 小于10的认为是费率，大于等于100的是固定金额门槛
                if 0 < num_val < 10 and num_val > max_fee:
                    max_fee = num_val
            except:
                pass

    return max_fee


def determine_fund_status(std_data):
    """判断基金状态：发行期 or 交易期"""
    issue_dates = std_data.get('发行日期', '')
    if not issue_dates:
        return '0:交易'

    start_date, end_date = extract_date_range(issue_dates)
    if not start_date:
        return '0:交易'

    today = datetime.now().strftime('%Y-%m-%d')

    # 如果募集结束日期 >= 今天，则处于发行期
    if end_date and end_date >= today:
        return '1:发行'
    elif start_date > today:
        return '1:发行'  # 还没开始募集
    else:
        return '0:交易'  # 募集已结束


def determine_has_subscription(std_data):
    """判断是否有认购业务"""
    status = determine_fund_status(std_data)
    # 募集期产品有认购业务
    if '1:发行' in status:
        return 'Y:是'
    else:
        return 'N:否'


def determine_detail_display_mode(std_data):
    """判断份额明细展示模式"""
    fund_name = std_data.get('基金简称', '')
    fund_full_name = std_data.get('基金名称', '')

    name_combined = str(fund_name) + str(fund_full_name)

    # 目标养老产品填养老模式（M）
    if '养老目标' in name_combined or '养老' in name_combined:
        return 'M:养老型模式'
    # 持有期产品（非养老）也填养老模式（M）
    elif '持有期' in name_combined:
        return 'M:养老型模式'
    # 滚动持有期产品
    elif '滚动' in name_combined:
        return 'G:滚动持有期'
    # 短期理财产品
    elif '短期理财' in name_combined:
        return 'Y:短期理财模式'
    else:
        return 'N:不展示'


def determine_investment_region(std_data):
    """判断投资地区"""
    fund_name = std_data.get('基金名称', '')
    fund_type = std_data.get('基金类型', '')

    fund_name = str(fund_name) if fund_name else ''
    fund_type = str(fund_type).upper() if fund_type else ''

    # QDII基金根据投资范围判断
    if 'QDII' in fund_type or 'QDII' in fund_name.upper():
        if '全球' in fund_name or '世界' in fund_name:
            return 'W:全球'
        elif '亚太' in fund_name or '亚洲' in fund_name:
            return 'A:亚太'
        elif '美国' in fund_name or '美股' in fund_name:
            return 'U:美国'
        elif '欧洲' in fund_name:
            return 'E:欧洲'
        elif '日本' in fund_name:
            return 'J:日本'
        elif '香港' in fund_name or '港股' in fund_name:
            return 'A:亚太'  # 香港属于亚太
        elif '拉美' in fund_name or '拉丁' in fund_name:
            return 'L:拉美'
        else:
            return 'W:全球'  # QDII默认全球

    # FOF产品默认亚太
    if 'FOF' in fund_type or 'FOF' in fund_name.upper() or '基金中基金' in fund_name:
        return 'A:亚太'

    # 非QDII/FOF基金默认中国大陆
    return 'M:中国大陆'


def determine_rolling_days(std_data):
    """判断滚动期开放天数"""
    fund_name = std_data.get('基金简称', '')

    if not fund_name:
        return None

    fund_name = str(fund_name)

    if '滚动' in fund_name:
        # 从名称中提取天数，如"30天滚动" -> 30
        match = re.search(r'(\d+)[天日]', fund_name)
        if match:
            return int(match.group(1))
        return 1  # 默认1天

    return None  # 非滚动持有期产品不用填写


def generate_param_mapping(std_data):
    """生成招行参数表字段映射值"""

    # 基本信息
    fund_code = std_data.get('基金代码', '')
    fund_name = std_data.get('基金名称', '')
    fund_short_name = std_data.get('基金简称', '')
    fund_type = std_data.get('基金类型', '')
    manager_code = std_data.get('基金管理人代码', '')
    custodian_bank = std_data.get('托管银行名称', '')
    manage_fee = std_data.get('管理费率（%）', 0)
    custodian_fee = std_data.get('托管费率（%）', 0)
    sales_fee = std_data.get('销售服务费率（%）', '')

    # 封闭期信息
    holding_period_num, holding_period_unit = extract_holding_period(str(fund_short_name))

    # 日期
    issue_dates = std_data.get('发行日期', '')
    start_date, end_date = extract_date_range(issue_dates)

    # 费率
    subscribe_fee = get_max_fee(std_data.get('基金认购费'))
    purchase_fee = get_max_fee(std_data.get('基金申购费'))
    redeem_fee_raw = std_data.get('赎回费率（%）', '')
    redeem_fee = parse_fee_value(redeem_fee_raw) if redeem_fee_raw else 0

    # 确认日/交收日
    subscribe_confirm = extract_tn_value(std_data.get('申购确认日（T+n）', ''))
    purchase_confirm = extract_tn_value(std_data.get('申购确认日（T+n）', ''))
    redeem_settle = extract_tn_value(std_data.get('赎回资金交收(净额)（T+n）', ''))
    dividend_settle = extract_tn_value(std_data.get('基金分红资金交收日期（R+n）', ''))

    # 账户信息
    subscribe_account = std_data.get('认购资金清算账户', {})
    purchase_account = std_data.get('申购资金清算账户', {})

    # 最低金额
    min_subscribe = std_data.get('最低认购投资金额（元）', 10)
    min_purchase = std_data.get('最低申购投资金额（元）', 1)
    min_add_subscribe = std_data.get('最低追加认购投资金额（元）', 10)
    min_add_purchase = std_data.get('最低追加申购投资金额（元）', 1)
    min_redeem = std_data.get('最低赎回份额（份）', 0.01)
    min_hold = std_data.get('最低保有份额（份）', 0.01)

    # 分红方式
    allow_div_mod = std_data.get('是否允许修改分红方式', '是')

    # 判断基金状态和是否有认购
    fund_status = determine_fund_status(std_data)
    has_subscription = determine_has_subscription(std_data)

    # 份额明细展示
    detail_display = determine_detail_display_mode(std_data)

    # 投资地区
    invest_region = determine_investment_region(std_data)

    # 滚动期天数
    rolling_days = determine_rolling_days(std_data)

    # 构建映射字典
    mapping = {
        # 产品基本信息
        'TA代码': f'F{manager_code}' if manager_code else 'F10',
        'SA代码': '007:招商银行',
        '基金代码': fund_code,
        '基金主代码': fund_code,
        '产品类型': 'A:开放式基金',
        '基金类型': get_fund_type_mapping(fund_type, fund_name),
        '显示模式': 'N:普通模式',
        '基金状态': fund_status,
        '投资地区': invest_region,
        '产品简称': fund_short_name,
        '基金长名称': fund_name,
        '封闭期参数': holding_period_num,
        '封闭期单位': f'{holding_period_unit}:{"个月" if holding_period_unit == "M" else "日" if holding_period_unit == "D" else "年" if holding_period_unit == "Y" else "季"}',
        '基金结算货币代码': '10:人民币',
        '基金结算货币市场': 'N:现钞',
        '托管机构名称': custodian_bank,
        '首次购买下限': min_subscribe,
        '基金管理费率(%)': parse_fee_value(manage_fee),
        '基金托管费率(%)': parse_fee_value(custodian_fee),
        '基金销售费率(%)': parse_fee_value(sales_fee) if sales_fee else 0,
        '默认收费方式': '0:前端收费',
        '回报速度(日)': purchase_confirm if purchase_confirm else 3,
        '净值日期更新频率': purchase_confirm if purchase_confirm else 3,
        '基金成立日期': '1900-01-01 00:00:00' if '1:发行' in fund_status else '',
        '份额明细展示': detail_display,
        '滚动期开放天数': rolling_days,
        '是否需要非美国居民声明': 'N:否',
        '养老目标基金风险揭示书': 'N:否',
        '是否上传基金产品资料概要': 'N:否',

        # 结算账户
        '结算方式': '3:非中登模式',
        '中登标准模式-认购': '2:全额非担保',
        '中登标准模式-申购': '1:全额担保',
        '中登标准模式-赎回': '1:全额担保',
        '中登标准模式-分红': '2:全额非担保',
        '是否有认购业务': has_subscription,
        '认购账户名称': subscribe_account.get('账户名称', ''),
        '认购账户开户银行': subscribe_account.get('开户银行', ''),
        '认购账户账号': str(subscribe_account.get('银行账号', '')).replace(' ', ''),
        '认购大额支付号/SWIFT CODE': str(subscribe_account.get('现代支付系统行号', '')).replace(' ', ''),
        '是否有申购业务': 'Y:是',
        '是否有申购业务申购资金交收方式': 'Y:净额交收',
        '申购账户名称': purchase_account.get('账户名称', ''),
        '申购账户开户银行': purchase_account.get('开户银行', ''),
        '申购账户账号': str(purchase_account.get('银行账号', '')).replace(' ', ''),
        '申购大额支付号/SWIFT CODE': str(purchase_account.get('现代支付系统行号', '')).replace(' ', ''),

        # 回款来源账户（自建TA模式）- 赎回款
        '赎回款-账户名称': purchase_account.get('账户名称', ''),
        '账号_赎回款': str(purchase_account.get('银行账号', '')).replace(' ', ''),
        '开户银行_赎回款': purchase_account.get('开户银行', ''),
        '账户标志_赎回款': 'X:非招行账户',
        '大额支付号/SWIFT CODE_赎回款': str(purchase_account.get('现代支付系统行号', '')).replace(' ', ''),

        # 分红款
        '分红款-账户名称': purchase_account.get('账户名称', ''),
        '账号_分红款': str(purchase_account.get('银行账号', '')).replace(' ', ''),
        '开户银行_分红款': purchase_account.get('开户银行', ''),
        '账户标志_分红款': 'X:非招行账户',
        '大额支付号/SWIFT CODE_分红款': str(purchase_account.get('现代支付系统行号', '')).replace(' ', ''),

        # 认购退款
        '认购退款-账户名称': subscribe_account.get('账户名称', ''),
        '账号_认购退款': str(subscribe_account.get('银行账号', '')).replace(' ', ''),
        '开户银行_认购退款': subscribe_account.get('开户银行', ''),
        '账户标志_认购退款': 'I:招行账户',
        '大额支付号/SWIFT CODE_认购退款': str(subscribe_account.get('现代支付系统行号', '')).replace(' ', ''),

        # 认购参数
        '认购总控': has_subscription,
        '认购资金扣款速度(日)': 1,
        '认购资金退款速度(日)': 0,
        '认购单位': 0.01,
        '认购每日上限': 99999999999.99,
        '认购单笔上限': 99999999999.99,
        '认购单笔下限': min_add_subscribe,
        '基金认购费率(展示)(%)': subscribe_fee,
        '募集开始日期': f'{start_date} 00:00:00' if start_date else '',
        '募集结束日期': f'{end_date} 00:00:00' if end_date else '',

        # 申购/定投
        '申购总控': 'Y:是',
        '申购资金清算速度(日)': 1,
        '申购单位': 0.01,
        '申购每日上限': 99999999999.99,
        '申购单笔上限': 99999999999.99,
        '申购单笔下限': min_add_purchase,
        '申购费率(%)': purchase_fee,
        '定投总控': 'Y:是',
        '定投资金清算速度(日)': 1,
        '定投投资单位': 0.01,
        '定投单笔上限': 99999999999.99,
        '定投单笔下限': std_data.get('最低定期定投金额（元）', 1),

        # 赎回/回款
        '赎回总控': 'Y:是',
        '分红方式修改': 'Y:是' if str(allow_div_mod) == '是' else 'N:否',
        '允许选择巨额赎回标志': 'Y:是',
        '赎回单位': 0.01,
        '赎回余额下限': min_hold,
        '赎回单笔上限': 99999999999.99,
        '赎回单笔下限': min_redeem,
        '赎回费率(%)': redeem_fee,
        '分红资金清算速度(日)': dividend_settle if dividend_settle else 6,
        '赎回/强制赎回资金到达银行的时间（T为业务申请日）': redeem_settle if redeem_settle else 6,

        # 基金转换
        '转换总控': 'N:否',
        '是否允许选择巨额赎回标志.1': 'N:否',
        '置换基数（单位）': 0.01,
        '置换单笔上限': 99999999999.99,
        '置换单笔下限': 0.01,
        '置换余额下限': 0,

        # 转托管
        '转托管转出总控': 'Y:是',
        '转出单笔上限': 99999999999.99,
        '转出单笔下限': 0.01,
        '转出余额下限': 0.01,
        '转入是否需要对方申请编号': 'M:必须',
    }

    return mapping


def fill_zhaohang_template(std_filepath, template_filepath, output_filepath):
    """填充招行参数表模板"""

    # 解析标准参数表
    std_data = parse_standard_param_table(std_filepath)

    # 生成映射值
    mapping = generate_param_mapping(std_data)

    # 加载招行参数表模板
    wb = openpyxl.load_workbook(template_filepath)
    ws = wb.active

    # 追踪当前section
    current_section = ''

    # 遍历A列，找到所有参数名，然后填充B列
    for row in range(1, ws.max_row + 1):
        param_name = ws.cell(row=row, column=1).value

        # 更新当前section
        if param_name in ['基金公司提供的产品相关参数', '结算账户', '回款来源账户信息（自建TA模式填写）', 
                           '认购参数', '申购,定投', '赎回,回款', '基金转换', '转托管']:
            current_section = param_name
            continue

        if param_name is None:
            continue

        param_name = str(param_name).strip()

        # 跳过标题行和空行
        if param_name in ['', '参数', '招商银行基金参数表']:
            continue

        # 获取映射值
        value = None

        # 特殊处理section内的重名字段
        if param_name in ['账号', '开户银行', '大额支付号/SWIFT CODE', '账户标志']:
            # 根据最近的父标题判断
            for r in range(row - 1, 0, -1):
                val = ws.cell(row=r, column=1).value
                if val and ('赎回款-' in str(val) or '分红款-' in str(val) or '认购退款-' in str(val)):
                    prefix = str(val).replace('-账户名称', '').replace('-', '_')
                    mapped_key = f"{param_name}{prefix}"
                    if mapped_key in mapping:
                        value = mapping[mapped_key]
                    break
                elif val and val in ['回款来源账户信息（自建TA模式填写）', '认购参数']:
                    break
        elif param_name == '是否允许选择巨额赎回标志':
            if current_section == '基金转换':
                value = mapping.get('是否允许选择巨额赎回标志.1')
            else:
                value = mapping.get(param_name)
        else:
            value = mapping.get(param_name)

        # 填充值
        if value is not None:
            ws.cell(row=row, column=2).value = value

    # 保存输出文件
    wb.save(output_filepath)
    return output_filepath


def main():
    if len(sys.argv) < 3:
        print("用法: python generate_zhaohang_param.py <标准参数表路径> <招行参数表模板路径> [输出路径]")
        sys.exit(1)

    std_filepath = sys.argv[1]
    template_filepath = sys.argv[2]

    if len(sys.argv) >= 4:
        output_filepath = sys.argv[3]
    else:
        # 自动生成输出文件名
        base_name = os.path.splitext(os.path.basename(template_filepath))[0]
        output_filepath = f'{base_name}_已填写.xlsx'

    result = fill_zhaohang_template(std_filepath, template_filepath, output_filepath)
    print(f"参数表已生成: {result}")


if __name__ == '__main__':
    main()
